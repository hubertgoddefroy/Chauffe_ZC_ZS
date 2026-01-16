import os
import time
import shutil
import subprocess
import threading  # <--- IMPORT IMPORTANT
from manipulation_image import *
from traitement_image import *
from openpyxl import Workbook
from manipulation_image import creation_dossier_snap


# Fonction pour lire la sortie en continu sans bloquer le script principal
def lire_sortie_continue(process, prefixe="[MACHINE]: "):
    """Lit stdout ligne par ligne tant que le process est en vie."""
    for ligne in iter(process.stdout.readline, ''):
        if ligne:
            print(f"{prefixe}{ligne.strip()}")
        else:
            break


def mesure_chauffe_ZC_snap(chemin_installation_ZS, chemin_dossier_enregistrement_image, nom_machine, puits, case_455,
                           case_730, case_455_730, prise_focus, nombre_iteration, frequence_enregistrement,
                           frequence_snap):
    # Lancement du processus
    process = subprocess.Popen(
        'cmd.exe',
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,  # On redirige aussi stderr pour voir les erreurs
        encoding='latin-1',
        errors='replace',
        text=True,
        shell=True,
        bufsize=0  # Tenter de désactiver le buffering (peut ne pas marcher sur tous les OS en mode texte)
    )

    # --- CORRECTION DU BLOCAGE ---
    # On lance un thread qui va "aspirer" tout ce que le programme sort.
    # Cela empêche le buffer de se remplir et de bloquer à l'itération 94.
    thread_lecture = threading.Thread(target=lire_sortie_continue, args=(process,))
    thread_lecture.daemon = True  # Le thread mourra quand le programme principal mourra
    thread_lecture.start()
    # -----------------------------

    # On écrit les commandes sans attendre de lire X lignes, car le thread le fait pour nous.
    process.stdin.write(
        chemin_installation_ZS + "/bin/ZymoCubeCtrl.exe " + chemin_installation_ZS + "/etc/ZymoCubeCtrl.ini" + '\n')
    process.stdin.flush()
    time.sleep(30)

    process.stdin.write(fr'reset AxHV_ZC_rfl_v1' + '\n')
    process.stdin.flush()
    # Plus besoin de lire manuellement ici, le thread affiche tout
    time.sleep(6)

    process.stdin.write(fr'start' + '\n')
    process.stdin.flush()
    time.sleep(6)

    process.stdin.write(fr'goto ' + puits + '\n')
    process.stdin.flush()
    time.sleep(7)

    ### Création fichier excel
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if case_455:
        onglet1 = wb.create_sheet(title="455")
    if case_730:
        onglet1 = wb.create_sheet(title="730")
    if case_455_730:
        onglet1 = wb.create_sheet(title="455")
        onglet2 = wb.create_sheet(title="730")

    # Initialisation de wb_temp et des onglets temporaires pour éviter les erreurs de référence avant la boucle
    wb_temp = None
    onglet1_temp = None
    onglet2_temp = None

    for i_boucle in range(int(nombre_iteration)):

        ### Création d'un fichier excel intermédiaire
        if i_boucle % int(frequence_enregistrement) == 0:
            wb_temp = Workbook()
            if "Sheet" in wb_temp.sheetnames:
                del wb_temp["Sheet"]

            if case_455:
                onglet1_temp = wb_temp.create_sheet(title="455")
            if case_730:
                onglet1_temp = wb_temp.create_sheet(title="730")
            if case_455_730:
                onglet1_temp = wb_temp.create_sheet(title="455")
                onglet2_temp = wb_temp.create_sheet(title="730")

        ### Prise des clichés
        if prise_focus:
            if case_455 or case_455_730:
                process.stdin.write(fr'goto ' + puits + '\n')
                process.stdin.flush()
                time.sleep(4)
                process.stdin.write(fr'snap 455' + '\n')
                process.stdin.flush()
                time.sleep(3)

            if case_730 or case_455_730:
                process.stdin.write(fr'goto ' + puits + '\n')
                process.stdin.flush()
                time.sleep(4)
                process.stdin.write(fr'snap 730' + '\n')
                process.stdin.flush()
                time.sleep(3)
        else:
            if case_455 or case_455_730:
                process.stdin.write(fr'snap 455' + '\n')
                process.stdin.flush()
                time.sleep(3)

            if case_730 or case_455_730:
                process.stdin.write(fr'snap 730' + '\n')
                process.stdin.flush()
                time.sleep(3)

        ### Récupération et traitement image (Le reste de votre logique reste identique)
        temperature_image_455 = 0
        intensite_image_455 = 0
        temperature_image_730 = 0
        intensite_image_730 = 0

        # NOTE: J'ai ajouté des try/except ici pour éviter un crash si l'image n'est pas encore là
        try:
            if case_455 or case_455_730:
                chemin_image_455 = obtenir_chemin_image_plus_recente_motif(chemin_dossier_enregistrement_image,
                                                                           puits + "_455", ".tif")
                temperature_image_455 = obtenir_temperature(chemin_image_455)
                intensite_image_455 = obtenir_intensite(25, 25, chemin_image_455)

            if case_730 or case_455_730:
                chemin_image_730 = obtenir_chemin_image_plus_recente_motif(chemin_dossier_enregistrement_image,
                                                                           puits + "_730", ".tif")
                temperature_image_730 = obtenir_temperature(chemin_image_730)
                intensite_image_730 = obtenir_intensite(25, 25, chemin_image_730)
        except Exception as e:
            print(f"Erreur lors de l'analyse d'image à l'itération {i_boucle}: {e}")
            # On continue quand même pour ne pas arrêter la boucle
            pass

        # Gestion de l'enregistrement des SNAP (Copier-coller)
        if i_boucle % int(frequence_snap) == 0:
            dossier_dest_base = os.path.join(chemin_dossier_enregistrement_image,
                                             "Snap_Chauffe_" + nom_machine + "_" + puits)
            if not os.path.exists(dossier_dest_base):
                os.makedirs(dossier_dest_base, exist_ok=True)

            if case_455 or case_455_730:
                try:
                    nom_fichier = f"{puits}_455({temperature_image_455})_{int(intensite_image_455)}_{i_boucle}.tif"
                    shutil.copy2(chemin_image_455, os.path.join(dossier_dest_base, nom_fichier))
                except Exception as e:
                    print(f"Erreur copie 455: {e}")

            if case_730 or case_455_730:
                try:
                    nom_fichier = f"{puits}_730({temperature_image_730})_{int(intensite_image_730)}_{i_boucle}.tif"
                    shutil.copy2(chemin_image_730, os.path.join(dossier_dest_base, nom_fichier))
                except Exception as e:
                    print(f"Erreur copie 730: {e}")

        ### Suppression des clichés pour éviter la saturation du stockage
        try:
            if case_455 or case_455_730:
                os.remove(chemin_image_455)
            if case_730 or case_455_730:
                os.remove(chemin_image_730)
        except Exception:
            pass  # Si le fichier n'existe pas ou est verrouillé, on ignore

        ### Remplissage du fichier excel intermédiaire
        if wb_temp:
            if case_455 and onglet1_temp:
                onglet1_temp.append([f"{i_boucle}", f"{temperature_image_455}", f"{intensite_image_455}"])
            if case_730 and onglet1_temp:
                onglet1_temp.append([f"{i_boucle}", f"{temperature_image_730}", f"{intensite_image_730}"])
            if case_455_730 and onglet1_temp and onglet2_temp:
                onglet1_temp.append([f"{i_boucle}", f"{temperature_image_455}", f"{intensite_image_455}"])
                onglet2_temp.append([f"{i_boucle}", f"{temperature_image_730}", f"{intensite_image_730}"])

        ### Remplissage du fichier excel final
        if case_455:
            onglet1.append([f"{i_boucle}", f"{temperature_image_455}", f"{intensite_image_455}"])
        if case_730:
            onglet1.append([f"{i_boucle}", f"{temperature_image_730}", f"{intensite_image_730}"])
        if case_455_730:
            onglet1.append([f"{i_boucle}", f"{temperature_image_455}", f"{intensite_image_455}"])
            onglet2.append([f"{i_boucle}", f"{temperature_image_730}", f"{intensite_image_730}"])

        ### Enregistrement du fichier excel intermédiaire
        if (i_boucle + 1) % int(frequence_enregistrement) == 0 and wb_temp:
            nom_temp = f"mesure_chauffe_{nom_machine}_{int((i_boucle + 1) / int(frequence_enregistrement))}.xlsx"
            try:
                wb_temp.save(os.path.join(chemin_dossier_enregistrement_image, nom_temp))
            except Exception as e:
                print(f"Erreur sauvegarde Excel temp: {e}")

        ### Attente
        # J'ai gardé vos temps, assurez-vous qu'ils sont suffisants pour la machine
        if prise_focus:
            if case_455_730:
                time.sleep(1)
            else:
                time.sleep(3.6)
        else:
            if case_455_730:
                time.sleep(4)
            else:
                time.sleep(6.6)

        print(f"Itération {i_boucle + 1} terminée")

    ### Enregistrement du fichier excel final
    try:
        wb.save(os.path.join(chemin_dossier_enregistrement_image, f"mesure_chauffe_{nom_machine}.xlsx"))
    except Exception as e:
        print(f"Erreur sauvegarde Excel final: {e}")

    process.stdin.write(fr'stop' + '\n')
    process.stdin.flush()
    print('STOPPED')
    time.sleep(10)
    process.stdin.write(fr'quit' + '\n')
    process.stdin.flush()
    time.sleep(3)
    print('EXITED')

    process.stdin.close()
    # On attend que le process se termine proprement
    process.wait()


def log_de_mesure(install_dir, save_dir, checkbox1_state, checkbox2_state, checkbox3_state, checkbox4_state,
                  text_field_machine,
                  text_field_iteration, text_field_frequence, combo_box_selection, text_field_frequence_snap):
    # (Cette fonction n'a pas besoin de changement, elle est OK)
    fichier_texte = save_dir + "/Calibration_" + text_field_machine + "_" + combo_box_selection + ".log"

    lignes = ["\n--- Données récupérées ---",
              "Dossier d'installation ZymoSoft : " + install_dir,
              "Dossier d'enregistrement des snap : " + save_dir,
              "455 ? " + str(checkbox1_state),
              "730 ? " + str(checkbox2_state),
              "455 & 730 ? " + str(checkbox3_state),
              "Nom Machine : " + text_field_machine,
              "Nombre d'itérations : " + text_field_iteration,
              "Période enregistrement du .csv : " + text_field_frequence,
              "Focus à chaque snap ? " + str(checkbox4_state),
              "Période enregistrement des snap : " + text_field_frequence_snap,
              "Puits acquis : " + combo_box_selection,
              "--------------------------\n"]

    with open(fichier_texte, 'w', encoding='utf-8') as fichier:
        for ligne in lignes:
            fichier.write(ligne + "\n")

    creation_dossier_snap(save_dir, "Snap_Chauffe_" + text_field_machine + "_" + combo_box_selection)